# NateMan – Nachschreibtermin-Manager
# exporter.py
# Copyright © 2020  Johannes Bingel
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from nateman import assigner
from .models import Klausur, Klausurteilnahme, Schueler, Stufe


def excelexport(filepath: str, stufen: Optional[Iterable[Stufe]] = None):
    # Intiziert die Exceldatei
    workbook = Workbook()

    # Holt sich alle Stufen falls keinen Angegeben sind
    if stufen is None:
        stufen = Stufe.query.all()

    # exstellt für jede Stufe ein Excelsheet
    for s in stufen:

        # Holt sich alle vNachschreiben in eine List aus der Datenbank
        k_schueler = Klausurteilnahme.query.join(Klausur).join(Schueler).filter(Klausurteilnahme.versaeumt) \
            .filter(~Klausurteilnahme.nachgeschrieben).filter(Klausur.stufe == s) \
            .order_by(Schueler.nachname, Schueler.vorname).all()

        # Holt sich die Zuordnugsvoschläge
        suggestions = assigner.zuordnen(k_schueler)

        # Initialisiert das Excelsheet
        workbook.create_sheet(s.name)
        worksheet = workbook[s.name]

        # Fügt die Schüle in die Exceltabelle ein.

        if len(k_schueler) != 0:

            worksheet.cell(1, 1, "Nr")
            worksheet.cell(1, 2, "Name")
            worksheet.cell(1, 3, "Vorname")
            worksheet.cell(1, 4, "Kurs")
            worksheet.cell(1, 5, "Lehrer")
            worksheet.cell(1, 6, "Dauer [min]")
            worksheet.cell(1, 7, "Attest")
            if suggestions is None:
                worksheet.cell(1, 8, "Bemerkung")
            else:
                worksheet.cell(1, 8, "Termin(Vorschlag)")
                worksheet.cell(1, 9, "Bemerkung")

            i = 0

            while i < len(k_schueler):
                worksheet.cell(i + 2, 1, i + 1)
                worksheet.cell(i + 2, 2, k_schueler[i].schueler.nachname)
                worksheet.cell(i + 2, 3, k_schueler[i].schueler.vorname)
                worksheet.cell(i + 2, 4, k_schueler[i].klausur.kursname)
                worksheet.cell(i + 2, 5, str(k_schueler[i].klausur.lehrer.kuerzel))
                worksheet.cell(i + 2, 6, k_schueler[i].klausur.laenge)
                if k_schueler[i].attestiert:
                    worksheet.cell(i + 2, 7, "Ja")
                else:
                    worksheet.cell(i + 2, 7, "Nein")

                nichtgefunden = True
                if suggestions is not None:
                    for i2 in range(0, len(suggestions)):
                        if k_schueler[i] in suggestions[i2]:
                            worksheet.cell(i + 2, 8, i2 + 1)
                            nichtgefunden = False
                    if nichtgefunden:
                        worksheet.cell(i + 2, 8, "problem")
                        nichtgefunden = True
                    worksheet.cell(i + 2, 9, k_schueler[i].klausur.annotation)
                else:
                    worksheet.cell(i + 2, 8, k_schueler[i].klausur.annotation)
                i += 1

            del i
            del nichtgefunden

        else:
            bold = Font(bold=True)
            worksheet.cell(1, 1, "Keine Schüler fehlen")
            worksheet.cell(1, 1, ).font = bold

    workbook.remove(workbook["Sheet"])
    workbook.close()
    workbook.save(filepath)
