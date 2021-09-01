# NateMan – Nachschreibtermin-Manager
# importer.py
# Copyright © 2020  Niklas Elsbrock und Johannes Bingel
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""
Enthält Funktionen zum Import von Daten.
"""
import io
import re
from datetime import datetime
from typing import TextIO, SupportsInt, Union

import bcrypt
from openpyxl import load_workbook
from sqlalchemy.sql import and_, exists
from xml.dom import minidom

from . import util
from .config_manager import config
from .models import Klausur, Klausurteilnahme, Lehrer, Schueler, Stufe, db, get_next_new_schueler_id


class KlausurplanImportError(Exception):
    """Beim Versuch, einen Klausurplan zu importieren aufgrund einer ungültigen Klausurdatei aufgetretener Fehler"""
    pass


class KoopSchuelerImportError(Exception):
    """Beim Versuch, Koop-Schueler zu importieren aufgrund einer ungültigen Klausurdatei aufgetretener Fehler"""
    pass


def _nbit_int(x: Union[str, bytes, SupportsInt], base=10, maxbits=32):
    result = int(x, base)
    if result.bit_length() > maxbits:
        raise ValueError(f"integer exceeds {maxbits}-bit limit")

    return result


def import_csv(file: TextIO, stufe: Stufe, new_lehrer_pwd: str) -> None:
    """
    Importiert eine Klausurliste aus der gegebenen CSV-Klausurplandatei für die angegebene Stufe.
    Registriert noch nicht registrierte Lehrer mit dem angegebenen Passwort.
    von Niklas Elsbrock

    :param file: zu importierende XML-Datei
    :param stufe: Stufe, für die der Plan importiert werden soll
    :param new_lehrer_pwd: Passwort für neu registrierte Lehrerkonten
    """
    col_kursname = 0
    col_lehrer_kuerzel = 1
    col_date = 2
    col_startperiod = 3
    col_endperiod = 4
    col_teilnehmer = 5

    row_len = 6

    delim_regex = r"[,;]"

    assert util.validate_bcrypt_password(new_lehrer_pwd)

    # Importdatum setzen
    stufe.import_date = datetime.now().date()

    new_lehrer_pwd_hash = bcrypt.hashpw(new_lehrer_pwd.encode("utf-8"), bcrypt.gensalt())

    rows = [tuple(re.split(delim_regex, row)) for row in io.TextIOWrapper(file).read().splitlines()]

    for rownum, row in enumerate(rows[1:], start=2):
        if len(row) != row_len:
            raise KlausurplanImportError(f"Zu importierender Plan für die Stufe {stufe.name} enthält in Zeile {rownum} "
                                         f"nur {len(row)} statt der erwarteten {row_len} Werte.")

        kursname = row[col_kursname]
        lehrer_kuerzel = row[col_lehrer_kuerzel]
        date = datetime.strptime(row[col_date], "%Y%m%d").date()
        startperiod = _nbit_int(row[col_startperiod])
        endperiod = _nbit_int(row[col_endperiod])

        if startperiod < 0 or startperiod > endperiod:
            raise KlausurplanImportError(f"Zu importierender Plan für die Stufe {stufe.name} enthält ungültigen "
                                         f"Zeitraum (Zeile: {rownum}; VonStd: {startperiod} BisStd: {endperiod})")

        # eventuell bereits registrierten Lehrer bekommen
        kurs_lehrer_entity = Lehrer.query.filter_by(kuerzel=lehrer_kuerzel).first()

        # Falls der Lehrer noch nicht registriert ist, registrieren
        if kurs_lehrer_entity is None:
            kurs_lehrer_entity = Lehrer(kuerzel=lehrer_kuerzel, is_confirmed=True, pwd_hash=new_lehrer_pwd_hash)
            # Falls ein E-Mailadressenformat konfiguriert ist, dieses nutzen.
            email_address_format = config.get("email-address-format", None)
            if email_address_format is not None:
                kurs_lehrer_entity.set_default_email_address(email_address_format)

            db.session.add(kurs_lehrer_entity)

        klausur_entity = Klausur(kursname=kursname, date=date, startperiod=startperiod, endperiod=endperiod,
                                 stufe=stufe, lehrer=kurs_lehrer_entity)
        db.session.add(klausur_entity)

        for schueler_entry in row[col_teilnehmer].split('~'):
            schueler_data = schueler_entry.rsplit('_', maxsplit=2)

            if len(schueler_data) != 3:
                schueler_entity = Schueler(id=get_next_new_schueler_id(), nachname=schueler_entry, vorname="???",
                                           stufe=stufe, koop=True)
            else:
                (nachname, vorname, birthdate) = schueler_data
                birthdate = int(birthdate)
                for rownum in range(100):
                    test_id = birthdate * 100 + rownum
                    test_schueler = Schueler.query.filter_by(id=test_id).first()
                    if test_schueler is None:
                        schueler_entity = Schueler(id=test_id, nachname=nachname, vorname=vorname, stufe=stufe)
                        break
                    elif test_schueler.nachname == nachname and test_schueler.vorname == vorname:
                        # Ist dieser Schüler bereits für eine andere Stufe eingetragen? -> Fehler
                        if test_schueler.stufe != stufe:
                            raise KlausurplanImportError(f"Zu importierender Plan für die {stufe.name} enthält Schüler "
                                                         f"({vorname} {nachname}, {birthdate}), der bereits für die "
                                                         f"{test_schueler.stufe.name} eingetragen ist "
                                                         f"(Zeile: {rownum}).")
                        schueler_entity = test_schueler
                        break
                else:
                    raise KlausurplanImportError(f"Zu importierender Plan für Stufe {stufe.name} enthält Schüler "
                                                 f"({vorname} {nachname}, {birthdate}; Zeile: {rownum}), dem keine ID "
                                                 f"gegeben werden konnte.")

            # Ist derselbe Schüler bereits in dieser Klausur? -> Fehler
            if db.session.query(exists().where(and_(Klausurteilnahme.schueler == schueler_entity,
                                                    Klausurteilnahme.klausur == klausur_entity))).scalar():
                raise KlausurplanImportError(f"Zu importierender Plan für die {stufe.name} enthält selben "
                                             f"Schüler ({schueler_entity.vorname} {schueler_entity.nachname}, "
                                             f"{schueler_entity.birthdate}, Zeile: {rownum}) mehrfach in derselben "
                                             f"Klausur ({kursname}).")

            # Schüler zur Klausur hinzufügen
            klausurteilnahme = Klausurteilnahme(klausur=klausur_entity, schueler=schueler_entity)
            db.session.add(klausurteilnahme)


def import_kurs42(xml_file: TextIO, stufe: Stufe, new_lehrer_pwd: str) -> None:
    """
    Importiert eine Klausurliste aus dem gegebenen XML-Klausurexport für die angegebene Stufe.
    Registriert noch nicht registrierte Lehrer mit dem angegebenen Passwort.
    von Niklas Elsbrock

    :param xml_file: zu importierende XML-Datei
    :param stufe: Stufe, für die die
    :param new_lehrer_pwd: Passwort für neu registrierte Lehrerkonten
    """
    assert util.validate_bcrypt_password(new_lehrer_pwd)

    # Importdatum setzen
    stufe.import_date = datetime.now().date()

    new_lehrer_pwd_hash = bcrypt.hashpw(new_lehrer_pwd.encode("utf-8"), bcrypt.gensalt())

    xmldoc = minidom.parse(xml_file)

    # Klausurschienen durchgehen
    for termin in xmldoc.getElementsByTagName("KLAUSURSCHIENE"):

        date = datetime.strptime(termin.attributes["Datum"].value, "%d.%m.%Y").date()
        startperiod = _nbit_int(termin.attributes["VonStd"].value)
        endperiod = _nbit_int(termin.attributes["BisStd"].value)

        if startperiod < 0 or startperiod > endperiod:
            raise KlausurplanImportError(f"Zu importierender Plan für die Stufe {stufe.name} enthält ungültigen "
                                         f"Zeitraum (Termin: {termin.attributes['Datum'].value}; VonStd: {startperiod} "
                                         f"BisStd: {endperiod})")

        if db.session.query(exists().where(and_(Klausur.date == date, Klausur.stufe == stufe))).scalar():
            raise KlausurplanImportError(f"Zu importierender Plan für die {stufe.name} enthält zwei Klausurtermine "
                                         f"(Klausurschienen) für dasselbe Datum.")

        # Kurse dieser Schiene durchgehen
        for kurs in termin.getElementsByTagName("KURS"):
            kursname = kurs.attributes["Bez"].value
            kurs_lehrerkuerzel = kurs.attributes["Lehrer"].value

            # eventuell bereits registrierten Lehrer bekommen
            kurs_lehrer_entity = Lehrer.query.filter_by(kuerzel=kurs_lehrerkuerzel).first()

            # Falls der Lehrer noch nicht registriert ist, registrieren
            if kurs_lehrer_entity is None:
                kurs_lehrer_entity = Lehrer(kuerzel=kurs_lehrerkuerzel, is_confirmed=True, pwd_hash=new_lehrer_pwd_hash)
                # Falls ein E-Mailadressenformat konfiguriert ist, dieses nutzen.
                email_address_format = config.get("email-address-format", None)
                if email_address_format is not None:
                    kurs_lehrer_entity.set_default_email_address(email_address_format)

                db.session.add(kurs_lehrer_entity)

            klausur_entity = Klausur(kursname=kursname, date=date, startperiod=startperiod, endperiod=endperiod,
                                     stufe=stufe, lehrer=kurs_lehrer_entity)
            db.session.add(klausur_entity)

            # Schüler dieses Kurses durchgehen
            for schueler in kurs.getElementsByTagName("SCHUELER"):
                schueler_dbid = _nbit_int(schueler.attributes["DbIdnr"].value)

                schueler_nachname = schueler.attributes["Name"].value.strip()
                schueler_vorname = schueler.attributes["Vorname"].value.strip()

                # eventuell bereits eingetragenen Schüler bekommen
                schueler_entity = Schueler.query.filter_by(id=schueler_dbid).first()

                # Falls der Schüler noch nicht eingetragen ist, eintragen
                if schueler_entity is None:
                    schueler_entity = Schueler(id=schueler_dbid, nachname=schueler_nachname,
                                               vorname=schueler_vorname, stufe=stufe)
                    db.session.add(schueler_entity)

                else:
                    # Ist dieser Schüler bereits für eine andere Stufe eingetragen? -> Fehler
                    if schueler_entity.stufe != stufe:
                        raise KlausurplanImportError(f"Zu importierender Plan für die {stufe.name} enthält Schüler "
                                                     f"({schueler_vorname} {schueler_nachname}, "
                                                     f"DbIdNr: {schueler_dbid}), der bereits für die "
                                                     f"{schueler_entity.stufe.name} eingetragen ist.")

                    # Ist derselbe Schüler bereits in dieser Klausur? -> Fehler
                    if db.session.query(exists().where(and_(Klausurteilnahme.schueler == schueler_entity,
                                                            Klausurteilnahme.klausur == klausur_entity))).scalar():
                        raise KlausurplanImportError(f"Zu importierender Plan für die {stufe.name} enthält selben "
                                                     f"Schüler ({schueler_vorname} {schueler_nachname}, "
                                                     f"DbIdNr: {schueler_dbid}) mehrfach in "
                                                     f"derselben Klausur ({kursname}).")

                # Falls der Schüler die Klausur mitschreibt, zur Klausur hinzufügen
                if schueler.attributes["Klausurschreiber"].value == "j":
                    klausurteilnahme = Klausurteilnahme(klausur=klausur_entity, schueler=schueler_entity)
                    db.session.add(klausurteilnahme)


# Von Johannes
def import_excel_koop(filepath: str):
    # Initialisiert Startdaten
    if Stufe.import_date is not None:
        workbook = load_workbook(filepath)
        worksheet = workbook["Datenblatt"]
        schueler = []
        exceptions = []

        # Holt die Daten der Schüler aus der Exceltabelle
        i = 10

        while worksheet.cell(row=i, column=3).value is not None and worksheet.cell(row=i + 1,
                                                                                   column=3).value is not None:
            if worksheet.cell(row=i, column=3).value is not None:
                schueler.append([worksheet.cell(row=i, column=6).value, worksheet.cell(row=i, column=3).value,
                                 worksheet.cell(row=i, column=4).value])
            i += 1

        # Spalten die Daten in ihre Bestandteile
        i = 0
        i2 = 0
        i3 = 0

        while i < len(schueler):
            while i2 < len(schueler[i][0]):
                if schueler[i][0][i2] == ",":
                    break
                i2 += 1
            while i3 < len(schueler[i][1]):
                if schueler[i][1][i3] == " ":
                    break
                i3 += 1
            schueler[i] = [schueler[i][0][:i2], schueler[i][0][i2 + 2:], schueler[i][1][:i3], schueler[i][1][i3 + 2:-1],
                           schueler[i][2]]
            i2 = 0
            i3 = 0
            i += 1

        # s 0 Nachname, 1 Vorname, 2 Fach , 3 Lehrer, 4 KOOP_Schuhle
        # Erstellt für jeden schüler einen Eintrag und prüft ob die Lks und Koopschulen existieren
        for s in schueler:
            if s[2] == "E":
                s[2] = "E5"
            klausuren = Klausur.query.join(Lehrer).filter(Lehrer.kuerzel == s[3]).all()
            i = 0
            while i < len(klausuren):
                if klausuren[i].kursname[:-2] in ((str(s[2]) + suffix) for suffix in ("-L", " L")):
                    i += 1
                else:
                    klausuren.pop(i)
            if len(klausuren) > 0:
                schueler_entity = Schueler(id=get_next_new_schueler_id(), nachname=s[0], vorname=s[1],
                                           stufe=klausuren[0].stufe, koop=True)
                for klausur in klausuren:
                    teilnahme_entity = Klausurteilnahme(klausur=klausur, schueler=schueler_entity)
                    db.session.add(schueler_entity)
                    db.session.add(teilnahme_entity)
            else:
                exceptions.append(s)

        return exceptions
    else:
        raise KoopSchuelerImportError("Für die Stufe ist noch kein Klausurplan importiert")
